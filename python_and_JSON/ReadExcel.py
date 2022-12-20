import openpyxl

# CARGAR ARCHIVO
wk = openpyxl.load_workbook("D:\\prueba.xlsx")

# SELECCIONAR HOJA DEL ARCHIVO
sh = wk['Hoja1']
print(sh.title)

# MOSTRAR VALORES DE CELDAS
print(sh['A2'].value)
celda = sh.cell(2,1)
celda2 = sh.cell(column=1, row=2)
print(celda.value)
print(celda2.value)

# LEER TODAS LAS FILAS Y COLUMNAS
rows = sh.max_row
columns = sh.max_column

print("\nMAX FILAS = " + str(rows) + " Y MAX COLUMNAS = " + str(columns))
print("\nMETODO CON RANGE:\n")
for i in range(1, rows+1):
    for j in range(1, columns+1):
        c=sh.cell(i,j)
        print(c.value)

print("\nMETODO CON sh:\n")
for r in sh['A1':'C3']:
    for c in r:
        print(c.value)