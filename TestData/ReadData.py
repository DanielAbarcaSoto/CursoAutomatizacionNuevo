import openpyxl

wk=openpyxl.load_workbook("D:\\Cursos BYTE\\robot framework\\TestData100.xlsx")

def fetch_number_of_rows(sheetname):
    sh = wk[sheetname]
    return sh.max_row

def fetch_cell_data(sheetname, row, col):
    sh = wk[sheetname]
    cell = sh.cell(int(row),int(col))
    return cell.value

